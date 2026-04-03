"""
Learner Tien Database Updater
==============================
Run this script to add new matches to LearnerTien_Database.xlsx
Usage: python update_tien_db.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os
import sys

# ── Config ────────────────────────────────────────────────────────
DB_FILE = "LearnerTien_Database.xlsx"

# Colors
GREEN_WIN  = "D6F0DA"
RED_LOSS   = "FFDADA"
WHITE      = "FFFFFF"
HARD_CLR   = "D6E4F7"
CLAY_CLR   = "FAE5CE"
GRASS_CLR  = "D6F5DC"
PALE_GREY  = "F4F6FA"

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def style_match_row(ws, row_num, result, surface):
    bg = GREEN_WIN if result == "W" else (RED_LOSS if result == "L" else WHITE)
    surf_bg = {"Hard": HARD_CLR, "Clay": CLAY_CLR, "Grass": GRASS_CLR}.get(surface, WHITE)

    for col in range(1, 19):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = PatternFill("solid", start_color=bg)
        cell.font = Font(name="Arial", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()
    # Surface column (col 3) gets surface color
    ws.cell(row=row_num, column=3).fill = PatternFill("solid", start_color=surf_bg)
    ws.row_dimensions[row_num].height = 16

def style_h2h_row(ws, row_num, win_pct_str, row_index):
    try:
        pct = float(win_pct_str.replace("%", ""))
        if pct >= 70:   bg = GREEN_WIN
        elif pct <= 30: bg = RED_LOSS
        else:           bg = PALE_GREY if row_index % 2 == 0 else WHITE
    except:
        bg = PALE_GREY if row_index % 2 == 0 else WHITE

    for col in range(1, 17):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = PatternFill("solid", start_color=bg)
        cell.font = Font(name="Arial", size=9, bold=(col == 1))
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row_num].height = 16

def get_input(prompt, options=None, optional=False):
    while True:
        val = input(f"  {prompt}: ").strip()
        if not val and optional:
            return ""
        if not val:
            print("    ⚠  This field is required.")
            continue
        if options and val not in options:
            print(f"    ⚠  Please enter one of: {', '.join(options)}")
            continue
        return val

def find_next_empty_row(ws, start_row=3, col=1):
    row = start_row
    while ws.cell(row=row, column=col).value is not None:
        row += 1
    return row

def update_h2h(ws_h2h, opponent, country, result):
    """Find opponent in H2H sheet and update their record, or add new row."""
    # Search for opponent
    found_row = None
    last_data_row = 2
    for row in ws_h2h.iter_rows(min_row=3):
        if row[0].value:
            last_data_row = row[0].row
            if str(row[0].value).strip().lower() == opponent.strip().lower():
                found_row = row[0].row
                break

    if found_row:
        # Update existing record
        r = found_row
        current_w = int(ws_h2h.cell(row=r, column=4).value or 0)
        current_l = int(ws_h2h.cell(row=r, column=5).value or 0)
        total = int(ws_h2h.cell(row=r, column=3).value or 0)

        if result == "W":
            current_w += 1
        else:
            current_l += 1
        total += 1

        win_pct = f"{(current_w/total)*100:.1f}%" if total > 0 else "-"

        ws_h2h.cell(row=r, column=3).value = total
        ws_h2h.cell(row=r, column=4).value = current_w
        ws_h2h.cell(row=r, column=5).value = current_l
        ws_h2h.cell(row=r, column=6).value = win_pct
        ws_h2h.cell(row=r, column=12).value = datetime.now().strftime("%d-%b-%Y")

        # Re-style the row
        row_index = r - 3
        style_h2h_row(ws_h2h, r, win_pct, row_index)
        print(f"    ✅ H2H updated: {opponent} → {current_w}W-{current_l}L ({win_pct})")

    else:
        # Add new opponent
        new_row = last_data_row + 1
        w = 1 if result == "W" else 0
        l = 1 if result == "L" else 0
        win_pct = "100.0%" if result == "W" else "0.0%"
        today = datetime.now().strftime("%d-%b-%Y")

        values = [opponent, country, 1, w, l, win_pct, 0, 0, 0, "-", today, today, "-", "-", "-", "-"]
        for ci, val in enumerate(values, 1):
            ws_h2h.cell(row=new_row, column=ci).value = val

        row_index = new_row - 3
        style_h2h_row(ws_h2h, new_row, win_pct, row_index)
        print(f"    ✅ H2H new entry added: {opponent}")

def add_match():
    print("\n" + "═"*55)
    print("  LEARNER TIEN DATABASE — NEW MATCH ENTRY")
    print("═"*55)

    if not os.path.exists(DB_FILE):
        print(f"\n  ❌ Error: '{DB_FILE}' not found.")
        print(f"     Make sure this script is in the same folder as {DB_FILE}")
        sys.exit(1)

    wb = openpyxl.load_workbook(DB_FILE)
    ws_matches = wb["📋 Match Log"]
    ws_h2h     = wb["🤝 H2H Records"]
    ws_ranking = wb["📈 Ranking History"]

    print("\n── Match Details ──────────────────────────────────")

    date       = get_input("Date (e.g. 15-Apr-2026)")
    tournament = get_input("Tournament name")
    surface    = get_input("Surface", options=["Hard", "Clay", "Grass"])
    level      = get_input("Level", options=["Grand Slam", "Masters 1000", "ATP 500", "ATP 250", "Challenger", "Special", "ITF"])
    round_     = get_input("Round (e.g. R128, R64, R32, R16, QF, SF, F)")
    tien_rank  = get_input("Tien's ranking at time")
    opp_rank   = get_input("Opponent ranking at time")
    opponent   = get_input("Opponent full name")
    country    = get_input("Opponent country code (e.g. USA, ITA, RUS)")
    result     = get_input("Result", options=["W", "L"])
    score      = get_input("Score (e.g. 6-3 7-5)")

    print("\n── Serve Stats (press Enter to skip) ──────────────")
    dr     = get_input("DR (Dominance Ratio, e.g. 1.15)", optional=True)
    ace    = get_input("Ace% (e.g. 8.2%)", optional=True)
    df     = get_input("DF% (e.g. 3.1%)", optional=True)
    first  = get_input("1st Serve In% (e.g. 62.0%)", optional=True)
    firstw = get_input("1st Serve Won% (e.g. 71.0%)", optional=True)
    secondw= get_input("2nd Serve Won% (e.g. 52.0%)", optional=True)
    bpsaved= get_input("BP Saved (e.g. 4/6)", optional=True)

    # ── Append to Match Log ──
    next_row = find_next_empty_row(ws_matches, start_row=3)
    row_data = [date, tournament, surface, level, round_,
                tien_rank, opp_rank, opponent, country, result, score,
                dr, ace, df, first, firstw, secondw, bpsaved]
    for ci, val in enumerate(row_data, 1):
        ws_matches.cell(row=next_row, column=ci).value = val

    style_match_row(ws_matches, next_row, result, surface)
    print(f"\n    ✅ Match appended to row {next_row} of Match Log")

    # ── Update H2H ──
    print("\n── Updating H2H Records ────────────────────────────")
    update_h2h(ws_h2h, opponent, country, result)

    # ── Save ──
    wb.save(DB_FILE)
    print(f"\n  💾 Saved → {DB_FILE}")
    print("═"*55 + "\n")

def add_multiple():
    another = True
    while another:
        add_match()
        cont = input("  Add another match? (y/n): ").strip().lower()
        another = (cont == "y")

def main():
    print("\n  LEARNER TIEN DATABASE UPDATER")
    print("  ─────────────────────────────")
    print("  1. Add a single new match")
    print("  2. Add multiple matches")
    print("  3. Exit")
    choice = input("\n  Choose (1/2/3): ").strip()

    if choice == "1":
        add_match()
    elif choice == "2":
        add_multiple()
    elif choice == "3":
        print("\n  Exiting.\n")
        sys.exit(0)
    else:
        print("  Invalid choice.")
        main()

if __name__ == "__main__":
    main()
